import openpyxl
import requests
import re
import json

from bs4 import BeautifulSoup


class Extract:
    """
    This class is used to extract data from a given url
    """
    def extract_data_from_url(self, url: str):
        text = ""
        html_text = requests.get(url).text
        parser = BeautifulSoup(html_text, "html.parser")
        links = [
            link.get("href")
            for link in parser.find_all("a", attrs={"href": re.compile("^https://")})
        ]

        images = [
            img.get("src")
            for img in parser.find_all("img", attrs={"src": re.compile(".png")})
        ]

        try:
            text = parser.find("body").text
            text = re.sub("[\n]+", "\n", text)  # extracting text from body tag of html
        except AttributeError:
            text = ""
        return links, images, text

worksheet = openpyxl.load_workbook("Scrapping.xlsx")
obj = worksheet.active
res = []
max_row = obj.max_row
max_col = obj.max_column
print(max_col, max_row)

all_links = [
    str(obj.cell(row=i, column=j).value)
    for i in range(1, max_row + 1)
    for j in range(1, max_col + 1)
]


# extracting urls , images and text from each url given in the excell sheet 
# and storing the data in a python dictionary
for link in all_links:
    links, images, text = Extract().extract_data_from_url(link)
    res.append({"WebPageLink": link, "urls": links, "images": images, "text": text})
res = {"data": res}

# converting and storing data in the form of a json file
with open("res_file.json", "w", encoding="utf-8") as f:
    json.dump(res, f)
